# ruff: noqa: E501
"""Safe parsing, validation, and persistence for the approved historical layout."""

from __future__ import annotations

import json
import re
import unicodedata
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from hashlib import sha256
from pathlib import Path
from typing import Any, Literal, cast

from openpyxl import load_workbook  # type: ignore[import-untyped]
from sqlalchemy import insert, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from lvfi_api.persistence.historical_models import (
    competitions,
    import_batches,
    import_issues,
    match_statistics,
    matches,
    seasons,
    source_records,
    teams,
)

APPROVED_SOURCE_SHA256 = (
    "93AF701AEF942A7C99004F7D95D8BE9D4DEEC81D07A260E376AF8E4ABED4FB7C"
)
HISTORICAL_HEADERS = (
    "Data",
    "Campeonato",
    "Temporada",
    "Mandante",
    "Visitante",
    "Gols_Mand_1T",
    "Gols_Vis_1T",
    "Gols_Mand_Jogo",
    "Gols_Vis_Jogo",
    "Fin_Mand_1T",
    "Fin_Vis_1T",
    "Fin_Mand_Jogo",
    "Fin_Vis_Jogo",
    "Chu_gol_Mand_1T",
    "Chu_gol_Vis_1T",
    "Chu_gol_Mand_Jogo",
    "Chu_gol_Vis_Jogo",
    "Esc_Mand_1T",
    "Esc_Vis_1T",
    "Esc_Mand_Jogo",
    "Esc_Vis_Jogo",
    "Fal_Mand_Jogo",
    "Fal_Vis_Jogo",
    "Cart_Mand_Jogo",
    "Cart_Vis_Jogo",
)
STATISTIC_FIELDS = (
    "home_goals_first_half",
    "away_goals_first_half",
    "home_goals_full_match",
    "away_goals_full_match",
    "home_shots_first_half",
    "away_shots_first_half",
    "home_shots_full_match",
    "away_shots_full_match",
    "home_shots_on_target_first_half",
    "away_shots_on_target_first_half",
    "home_shots_on_target_full_match",
    "away_shots_on_target_full_match",
    "home_corners_first_half",
    "away_corners_first_half",
    "home_corners_full_match",
    "away_corners_full_match",
    "home_fouls_full_match",
    "away_fouls_full_match",
    "home_cards_full_match",
    "away_cards_full_match",
)
FIRST_HALF_PAIRS = (
    (0, 2),
    (1, 3),
    (4, 6),
    (5, 7),
    (8, 10),
    (9, 11),
    (12, 14),
    (13, 15),
)
TARGET_SHOT_PAIRS = ((8, 4), (9, 5), (10, 6), (11, 7))
INTEGER_TEXT = re.compile(r"^[+-]?\d+$")


class SourceValidationError(ValueError):
    """Raised when a workbook is not an approved structural source."""


@dataclass(frozen=True)
class ImportIssue:
    """A row-level validation result retained independently from accepted data."""

    severity: Literal["error", "warning"]
    code: str
    field_name: str | None
    message: str


@dataclass(frozen=True)
class NormalizedRow:
    """A valid, normalized representation of one source row."""

    source_line: int
    raw_values: dict[str, Any]
    row_sha256: str
    played_on: date
    competition: str
    competition_key: str
    season: str
    home_team: str
    home_team_key: str
    away_team: str
    away_team_key: str
    statistics: dict[str, int]


@dataclass(frozen=True)
class ImportSummary:
    """Aggregated outcome without exposing proprietary source rows."""

    source_sha256: str
    sheet_name: str
    total_records: int
    accepted_records: int
    rejected_records: int
    warning_records: int
    dry_run: bool
    already_imported: bool = False


def normalized_key(value: str) -> str:
    """Return the conservative key used only for exact identity lookup."""

    return unicodedata.normalize("NFKC", value).strip().casefold()


def source_sha256(path: Path) -> str:
    """Hash source bytes without changing the file."""

    digest = sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _raw_json_value(value: object) -> object:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def row_sha256(raw_values: dict[str, Any]) -> str:
    """Create a deterministic identity from the complete raw row representation."""

    payload = json.dumps(
        raw_values, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    )
    return sha256(payload.encode("utf-8")).hexdigest()


def _required_text(
    value: object, field_name: str, issues: list[ImportIssue]
) -> tuple[str, str] | None:
    if not isinstance(value, str) or not value.strip():
        issues.append(
            ImportIssue(
                "error", "required_text", field_name, "required text is missing"
            )
        )
        return None
    displayed = unicodedata.normalize("NFKC", value).strip()
    return displayed, normalized_key(displayed)


def _date_value(value: object, issues: list[ImportIssue]) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for pattern in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(value.strip(), pattern).date()
            except ValueError:
                pass
    issues.append(
        ImportIssue("error", "invalid_date", "Data", "date is missing or invalid")
    )
    return None


def _integer_value(
    value: object, field_name: str, issues: list[ImportIssue]
) -> int | None:
    parsed: int | None
    if isinstance(value, int) and not isinstance(value, bool):
        parsed = value
    elif isinstance(value, str) and INTEGER_TEXT.fullmatch(value.strip()):
        parsed = int(value.strip())
    else:
        issues.append(
            ImportIssue(
                "error", "invalid_integer", field_name, "statistic is not an integer"
            )
        )
        return None
    if parsed < 0:
        issues.append(
            ImportIssue(
                "error",
                "negative_statistic",
                field_name,
                "statistic cannot be negative",
            )
        )
        return None
    return parsed


def normalize_row(
    values: Sequence[object], source_line: int
) -> tuple[NormalizedRow | None, tuple[ImportIssue, ...]]:
    """Validate one exact-layout row without filling or correcting source values."""

    issues: list[ImportIssue] = []
    if len(values) != len(HISTORICAL_HEADERS):
        issue = ImportIssue(
            "error", "incomplete_row", None, "row does not have 25 columns"
        )
        return None, (issue,)
    raw_values = {
        header: _raw_json_value(value)
        for header, value in zip(HISTORICAL_HEADERS, values, strict=True)
    }
    played_on = _date_value(values[0], issues)
    competition = _required_text(values[1], "Campeonato", issues)
    season = _required_text(
        str(values[2]) if values[2] is not None else None, "Temporada", issues
    )
    home = _required_text(values[3], "Mandante", issues)
    away = _required_text(values[4], "Visitante", issues)
    statistics = {
        field_name: _integer_value(value, header, issues)
        for field_name, header, value in zip(
            STATISTIC_FIELDS, HISTORICAL_HEADERS[5:], values[5:], strict=True
        )
    }
    if home is not None and away is not None and home[1] == away[1]:
        issues.append(
            ImportIssue("error", "same_teams", None, "home and away teams must differ")
        )
    if all(isinstance(statistics[field], int) for field in STATISTIC_FIELDS):
        checked = [cast(int, statistics[field]) for field in STATISTIC_FIELDS]
        for first_half, full_match in FIRST_HALF_PAIRS:
            if checked[first_half] > checked[full_match]:
                issues.append(
                    ImportIssue(
                        "error",
                        "first_half_exceeds_full",
                        HISTORICAL_HEADERS[5 + first_half],
                        "first-half statistic exceeds full-match statistic",
                    )
                )
        for targets, shots in TARGET_SHOT_PAIRS:
            if checked[targets] > checked[shots]:
                issues.append(
                    ImportIssue(
                        "error",
                        "shots_on_target_exceed_shots",
                        HISTORICAL_HEADERS[5 + targets],
                        "shots on target exceed shots",
                    )
                )
    if issues:
        return None, tuple(issues)
    assert (
        played_on is not None
        and competition is not None
        and season is not None
        and home is not None
        and away is not None
    )
    return NormalizedRow(
        source_line=source_line,
        raw_values=raw_values,
        row_sha256=row_sha256(raw_values),
        played_on=played_on,
        competition=competition[0],
        competition_key=competition[1],
        season=season[0],
        home_team=home[0],
        home_team_key=home[1],
        away_team=away[0],
        away_team_key=away[1],
        statistics={
            field: value for field, value in statistics.items() if value is not None
        },
    ), tuple()


def read_source(
    path: Path, sheet_name: str, expected_sha256: str | None = None
) -> tuple[str, Iterable[tuple[int, tuple[object, ...]]]]:
    """Verify an XLSM source and return a streaming row iterator in safe read mode."""

    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise SourceValidationError("source must be an .xlsx or .xlsm file")
    actual_hash = source_sha256(path)
    if expected_sha256 is not None and actual_hash != expected_sha256.upper():
        raise SourceValidationError("source SHA-256 does not match the approved value")
    workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise SourceValidationError("expected sheet is missing")
    worksheet = workbook[sheet_name]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    if tuple(header_row) != HISTORICAL_HEADERS:
        workbook.close()
        raise SourceValidationError(
            "source headers do not match the historical contract"
        )

    def rows() -> Iterable[tuple[int, tuple[object, ...]]]:
        try:
            for line, row in enumerate(
                worksheet.iter_rows(min_row=2, max_col=25, values_only=True), start=2
            ):
                yield line, tuple(row)
        finally:
            workbook.close()
            if source_sha256(path) != actual_hash:
                raise SourceValidationError("source changed during safe reading")

    return actual_hash, rows()


class HistoricalImporter:
    """Persist validated source rows with idempotent batch and record identities."""

    def __init__(self, session: AsyncSession) -> None:
        self._session = session

    async def _id_for(
        self,
        table: Any,
        key_column: Any,
        key_value: str,
        display_name: str | None = None,
    ) -> int:
        existing = await self._session.scalar(
            select(table.c.id).where(key_column == key_value)
        )
        if existing is not None:
            return int(existing)
        values = {key_column.name: key_value}
        if display_name is not None:
            values["display_name"] = display_name
        return int(
            (
                await self._session.execute(
                    insert(table).values(**values).returning(table.c.id)
                )
            ).scalar_one()
        )

    async def _season_id(self, competition_id: int, label: str) -> int:
        existing = await self._session.scalar(
            select(seasons.c.id).where(
                seasons.c.competition_id == competition_id, seasons.c.label == label
            )
        )
        if existing is not None:
            return int(existing)
        return int(
            (
                await self._session.execute(
                    insert(seasons)
                    .values(competition_id=competition_id, label=label)
                    .returning(seasons.c.id)
                )
            ).scalar_one()
        )

    async def execute(
        self, path: Path, sheet_name: str, expected_sha256: str | None, dry_run: bool
    ) -> ImportSummary:
        """Validate, optionally persist, and summarize one controlled workbook import."""

        file_hash, rows = read_source(path, sheet_name, expected_sha256)
        parsed = [
            (
                line,
                {
                    header: _raw_json_value(value)
                    for header, value in zip(HISTORICAL_HEADERS, values, strict=True)
                },
                *normalize_row(values, line),
            )
            for line, values in rows
        ]
        accepted = sum(normalized is not None for _, _, normalized, _ in parsed)
        rejected = len(parsed) - accepted
        warnings = sum(
            any(issue.severity == "warning" for issue in issues)
            for _, _, _, issues in parsed
        )
        if dry_run:
            return ImportSummary(
                file_hash, sheet_name, len(parsed), accepted, rejected, warnings, True
            )
        batch_id = await self._session.scalar(
            select(import_batches.c.id).where(
                import_batches.c.source_sha256 == file_hash,
                import_batches.c.sheet_name == sheet_name,
            )
        )
        if batch_id is not None:
            return ImportSummary(
                file_hash,
                sheet_name,
                len(parsed),
                accepted,
                rejected,
                warnings,
                False,
                True,
            )
        batch_id = int(
            (
                await self._session.execute(
                    insert(import_batches)
                    .values(
                        source_filename=path.name,
                        source_sha256=file_hash,
                        sheet_name=sheet_name,
                        status="running",
                    )
                    .returning(import_batches.c.id)
                )
            ).scalar_one()
        )
        persisted_accepted = 0
        persisted_rejected = 0
        for line, raw_values, normalized, issues in parsed:
            digest = (
                normalized.row_sha256
                if normalized is not None
                else row_sha256(raw_values)
            )
            source_status = "accepted" if normalized is not None else "rejected"
            source_id = int(
                (
                    await self._session.execute(
                        insert(source_records)
                        .values(
                            batch_id=batch_id,
                            source_line=line,
                            row_sha256=digest,
                            raw_values=raw_values,
                            status=source_status,
                        )
                        .returning(source_records.c.id)
                    )
                ).scalar_one()
            )
            for issue in issues:
                await self._session.execute(
                    insert(import_issues).values(
                        source_record_id=source_id,
                        severity=issue.severity,
                        code=issue.code,
                        field_name=issue.field_name,
                        message=issue.message,
                    )
                )
            if normalized is None:
                persisted_rejected += 1
                continue
            competition_id = await self._id_for(
                competitions,
                competitions.c.normalized_name,
                normalized.competition_key,
                normalized.competition,
            )
            season_id = await self._season_id(competition_id, normalized.season)
            home_id = await self._id_for(
                teams,
                teams.c.normalized_name,
                normalized.home_team_key,
                normalized.home_team,
            )
            away_id = await self._id_for(
                teams,
                teams.c.normalized_name,
                normalized.away_team_key,
                normalized.away_team,
            )
            existing_match = await self._session.scalar(
                select(matches.c.id).where(
                    matches.c.season_id == season_id,
                    matches.c.played_on == normalized.played_on,
                    matches.c.home_team_id == home_id,
                    matches.c.away_team_id == away_id,
                )
            )
            if existing_match is not None:
                await self._session.execute(
                    update(source_records)
                    .where(source_records.c.id == source_id)
                    .values(status="conflict")
                )
                await self._session.execute(
                    insert(import_issues).values(
                        source_record_id=source_id,
                        severity="error",
                        code="canonical_match_conflict",
                        field_name=None,
                        message="canonical match key already exists",
                    )
                )
                persisted_rejected += 1
                continue
            match_id = int(
                (
                    await self._session.execute(
                        insert(matches)
                        .values(
                            season_id=season_id,
                            played_on=normalized.played_on,
                            home_team_id=home_id,
                            away_team_id=away_id,
                            source_record_id=source_id,
                        )
                        .returning(matches.c.id)
                    )
                ).scalar_one()
            )
            await self._session.execute(
                insert(match_statistics).values(
                    match_id=match_id, **normalized.statistics
                )
            )
            persisted_accepted += 1
        status = "completed" if persisted_rejected == 0 else "completed_with_rejections"
        await self._session.execute(
            update(import_batches)
            .where(import_batches.c.id == batch_id)
            .values(
                status=status,
                total_records=len(parsed),
                accepted_records=persisted_accepted,
                rejected_records=persisted_rejected,
                warning_records=warnings,
                completed_at=datetime.now().astimezone(),
            )
        )
        return ImportSummary(
            file_hash,
            sheet_name,
            len(parsed),
            persisted_accepted,
            persisted_rejected,
            warnings,
            False,
        )
